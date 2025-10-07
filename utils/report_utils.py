# report_utils.py - Utilities for reporting and data export
import os
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple
import csv
from .message_utils import error, success, warning, info

# Try to import pandas and openpyxl for enhanced Excel features
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ReportManager:
    
    def __init__(self, base_path: str = "reports"):
        self.base_path = base_path
        self.ensure_directories()
        
    def ensure_directories(self):
        dirs = [
            "benh_nhan", "tiep_nhan", "dich_vu", 
            "phong_kham", "tong_hop", "thong_ke"
        ]
        for dir_name in dirs:
            os.makedirs(os.path.join(self.base_path, dir_name), exist_ok=True)
    
    def generate_filename(self, report_type: str, file_format: str = "xlsx", 
                         prefix: str = "", suffix: str = "") -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if prefix:
            prefix = f"{prefix}_"
        if suffix:
            suffix = f"_{suffix}"
        return f"{prefix}{report_type}{suffix}_{timestamp}.{file_format}"
    
    def get_report_path(self, category: str, filename: str) -> str:
        return os.path.join(self.base_path, category, filename)
    
    def create_excel_with_styling(self, data: List[Dict], filename: str, 
                                 sheet_name: str = "Data", 
                                 title: str = "", author: str = "") -> str:
        if not EXCEL_AVAILABLE:
            return self.create_csv_fallback(data, filename.replace('.xlsx', '.csv'))
        
        try:
            os.makedirs(os.path.dirname(filename), exist_ok=True)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            so_dong_hien_tai = 1
            if title:
                ws.merge_cells(f"A{so_dong_hien_tai}:Z{so_dong_hien_tai}")
                title_cell = ws[f"A{so_dong_hien_tai}"]
                title_cell.value = title
                title_cell.font = Font(bold=True, size=16)
                title_cell.alignment = Alignment(horizontal="center")
                so_dong_hien_tai += 2
            
            if author:
                ws[f"A{so_dong_hien_tai}"] = f"Người xuất: {author}"
                so_dong_hien_tai += 1
            
            ws[f"A{so_dong_hien_tai}"] = f"Thời gian xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            so_dong_hien_tai += 2
            
            if data:
                headers = list(data[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=so_dong_hien_tai, column=col_idx)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                so_dong_hien_tai += 1
                
                for row_data in data:
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=so_dong_hien_tai, column=col_idx)
                        cell.value = row_data.get(header, "")
                        cell.border = thin_border
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right")
                    so_dong_hien_tai += 1
                
                for col_idx, column in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(col_idx)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            return filename
            
        except Exception as e:
            error(f"Lỗi tạo Excel: {e}")
            return self.create_csv_fallback(data, filename.replace('.xlsx', '.csv'))
    
    def create_csv_fallback(self, data: List[Dict], filename: str) -> str:
        try:
            os.makedirs(os.path.dirname(filename), exist_ok=True)
            
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                if data:
                    writer = csv.DictWriter(csvfile, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)
            return filename
        except Exception as e:
            error(f"Lỗi tạo CSV: {e}")
            return ""
    
    def create_summary_report(self, data_dict: Dict[str, List[Dict]], 
                            filename: str, title: str = "", author: str = "") -> str:
        if not EXCEL_AVAILABLE:
            for sheet_name, data in data_dict.items():
                csv_filename = filename.replace('.xlsx', f'_{sheet_name}.csv')
                os.makedirs(os.path.dirname(csv_filename), exist_ok=True)
                self.create_csv_fallback(data, csv_filename)
            return filename.replace('.xlsx', '_multiple_csv')
        
        try:
            os.makedirs(os.path.dirname(filename), exist_ok=True)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                summary_data = []
                for sheet_name, data in data_dict.items():
                    summary_data.append({
                        'Sheet': sheet_name,
                        'Số bản ghi': len(data),
                        'Cập nhật': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                    })
                
                if summary_data:
                    pd.DataFrame(summary_data).to_excel(
                        writer, sheet_name='Tổng hợp', index=False
                    )
                
                for sheet_name, data in data_dict.items():
                    if data:
                        df = pd.DataFrame(data)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            return filename
        except Exception as e:
            error(f"Lỗi tạo báo cáo tổng hợp: {e}")
            return ""

    def get_statistics_summary(self, model) -> Dict[str, Any]:
        try:
            stats = {}
            
            # Thống kê cơ bản
            stats['benh_nhan'] = {
                'tong_so': len(model.ds_benh_nhan()),
                'nam': 0,
                'nu': 0
            }
            
            for bn in model.ds_benh_nhan():
                if hasattr(bn, '_gioi_tinh'):
                    if bn._gioi_tinh.lower() in ['nam', 'male', 'm']:
                        stats['benh_nhan']['nam'] += 1
                    elif bn._gioi_tinh.lower() in ['nu', 'female', 'f', 'nữ']:
                        stats['benh_nhan']['nu'] += 1
            
            stats['tiep_nhan'] = {
                'tong_so': len(model.ds_tiep_nhan()),
                'hom_nay': 0,
                'tuan_nay': 0,
                'thang_nay': 0
            }
            
            now = datetime.now()
            today = now.date()
            week_start = today - timedelta(days=today.weekday())
            month_start = today.replace(day=1)
            
            for tn in model.ds_tiep_nhan():
                # Giả sử có thuộc tính ngay_tao
                if hasattr(tn, 'ngay_tao') and tn.ngay_tao:
                    tn_date = tn.ngay_tao.date() if hasattr(tn.ngay_tao, 'date') else tn.ngay_tao
                    if tn_date == today:
                        stats['tiep_nhan']['hom_nay'] += 1
                    if tn_date >= week_start:
                        stats['tiep_nhan']['tuan_nay'] += 1
                    if tn_date >= month_start:
                        stats['tiep_nhan']['thang_nay'] += 1
            
            stats['dich_vu'] = {
                'tong_so': len(model.ds_dich_vu()),
                'gia_cao_nhat': 0,
                'gia_thap_nhat': float('inf')
            }
            
            for dv in model.ds_dich_vu():
                if hasattr(dv, '_gia') and isinstance(dv._gia, (int, float)):
                    if dv._gia > stats['dich_vu']['gia_cao_nhat']:
                        stats['dich_vu']['gia_cao_nhat'] = dv._gia
                    if dv._gia < stats['dich_vu']['gia_thap_nhat']:
                        stats['dich_vu']['gia_thap_nhat'] = dv._gia
            
            if stats['dich_vu']['gia_thap_nhat'] == float('inf'):
                stats['dich_vu']['gia_thap_nhat'] = 0
                
            stats['phong_kham'] = {
                'tong_so': len(model.ds_phong_kham())
            }
            
            stats['bac_si'] = {
                'tong_so': len(model.ds_bac_si()) if hasattr(model, 'ds_bac_si') and callable(model.ds_bac_si) else 0
            }
            
            return stats
            
        except Exception as e:
            error(f"Lỗi khi lấy thống kê: {e}")
            return {}

    def export_bao_cao_thong_ke(self, model, author: str = "System") -> str:
        """Xuất báo cáo thống kê tổng hợp"""
        try:
            stats = self.get_statistics_summary(model)
            
            data_dict = {}
            
            overview_data = [
                {'Chỉ số': 'Tổng số bệnh nhân', 'Giá trị': stats.get('benh_nhan', {}).get('tong_so', 0)},
                {'Chỉ số': '- Bệnh nhân nam', 'Giá trị': stats.get('benh_nhan', {}).get('nam', 0)},
                {'Chỉ số': '- Bệnh nhân nữ', 'Giá trị': stats.get('benh_nhan', {}).get('nu', 0)},
                {'Chỉ số': 'Tổng số tiếp nhận', 'Giá trị': stats.get('tiep_nhan', {}).get('tong_so', 0)},
                {'Chỉ số': '- Tiếp nhận hôm nay', 'Giá trị': stats.get('tiep_nhan', {}).get('hom_nay', 0)},
                {'Chỉ số': '- Tiếp nhận tuần này', 'Giá trị': stats.get('tiep_nhan', {}).get('tuan_nay', 0)},
                {'Chỉ số': '- Tiếp nhận tháng này', 'Giá trị': stats.get('tiep_nhan', {}).get('thang_nay', 0)},
                {'Chỉ số': 'Tổng số dịch vụ', 'Giá trị': stats.get('dich_vu', {}).get('tong_so', 0)},
                {'Chỉ số': '- Giá cao nhất (VNĐ)', 'Giá trị': f"{stats.get('dich_vu', {}).get('gia_cao_nhat', 0):,}"},
                {'Chỉ số': '- Giá thấp nhất (VNĐ)', 'Giá trị': f"{stats.get('dich_vu', {}).get('gia_thap_nhat', 0):,}"},
                {'Chỉ số': 'Tổng số phòng khám', 'Giá trị': stats.get('phong_kham', {}).get('tong_so', 0)},
                {'Chỉ số': 'Tổng số bác sĩ', 'Giá trị': stats.get('bac_si', {}).get('tong_so', 0)},
            ]
            data_dict['Thống kê tổng quan'] = overview_data
            
            benh_nhan_data = []
            for i, bn in enumerate(model.ds_benh_nhan(), 1):
                benh_nhan_data.append({
                    'STT': i,
                    'Mã BN': getattr(bn, 'ma_bn', ''),
                    'PID': getattr(bn, 'pid', ''),
                    'Họ tên': getattr(bn, '_ho_ten', ''),
                    'Giới tính': getattr(bn, '_gioi_tinh', ''),
                    'Năm sinh': getattr(bn, '_nam_sinh', ''),
                    'CCCD': getattr(bn, '_so_cccd', '')
                })
            if benh_nhan_data:
                data_dict['Chi tiết bệnh nhân'] = benh_nhan_data
            
            tiep_nhan_data = []
            for i, tn in enumerate(model.ds_tiep_nhan(), 1):
                tiep_nhan_data.append({
                    'STT': i,
                    'Mã TN': getattr(tn, '_ma_tn', ''),
                    'Tên bệnh nhân': getattr(getattr(tn, '_bn', None), '_ho_ten', '') if hasattr(tn, '_bn') else '',
                    'CCCD': getattr(getattr(tn, '_bn', None), '_so_cccd', '') if hasattr(tn, '_bn') else '',
                    'Dịch vụ': getattr(getattr(tn, '_dv', None), '_ten_dv', '') if hasattr(tn, '_dv') else '',
                    'Phòng khám': getattr(getattr(tn, '_pk', None), '_ten_phong', '') if hasattr(tn, '_pk') else '',
                    'Lý do': getattr(tn, '_ly_do', ''),
                })
            if tiep_nhan_data:
                data_dict['Chi tiết tiếp nhận'] = tiep_nhan_data
            
            filename = self.generate_filename("ThongKe_TongHop", "xlsx", "BaoCao")
            filepath = self.get_report_path("thong_ke", filename)
            
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            result_path = self.create_summary_report(
                data_dict, 
                filepath, 
                "BÁO CÁO THỐNG KÊ TỔNG HỢP HỆ THỐNG", 
                author
            )
            
            return result_path
            
        except Exception as e:
            error(f"Lỗi khi xuất báo cáo thống kê: {e}")
            return ""

    def export_bao_cao_doanh_thu(self, model, author: str = "System") -> str:
        try:
            tiep_nhan_list = model.ds_tiep_nhan()
            
            revenue_data = []
            total_revenue = 0
            service_revenue = {}
            
            for tn in tiep_nhan_list:
                if hasattr(tn, '_dv') and tn._dv and hasattr(tn._dv, '_gia'):
                    service_name = getattr(tn._dv, '_ten_dv', 'Unknown')
                    service_price = getattr(tn._dv, '_gia', 0)
                    benh_nhan_name = getattr(getattr(tn, '_bn', None), '_ho_ten', '') if hasattr(tn, '_bn') else ''
                    
                    revenue_data.append({
                        'Mã TN': getattr(tn, '_ma_tn', ''),
                        'Bệnh nhân': benh_nhan_name,
                        'Dịch vụ': service_name,
                        'Giá (VNĐ)': service_price,
                        'Định dạng giá': f"{service_price:,}"
                    })
                    
                    total_revenue += service_price
                    
                    if service_name not in service_revenue:
                        service_revenue[service_name] = {'count': 0, 'total': 0}
                    service_revenue[service_name]['count'] += 1
                    service_revenue[service_name]['total'] += service_price
            
            data_dict = {}
            
            summary_data = [
                {'Chỉ số': 'Tổng doanh thu (VNĐ)', 'Giá trị': f"{total_revenue:,}"},
                {'Chỉ số': 'Số lần tiếp nhận', 'Giá trị': len(revenue_data)},
                {'Chỉ số': 'Doanh thu trung bình/lần', 'Giá trị': f"{total_revenue // len(revenue_data) if revenue_data else 0:,}"},
                {'Chỉ số': 'Số loại dịch vụ được sử dụng', 'Giá trị': len(service_revenue)},
            ]
            data_dict['Tổng quan doanh thu'] = summary_data
            
            if revenue_data:
                data_dict['Chi tiết doanh thu'] = revenue_data
            
            service_stats = []
            for service_name, stats in service_revenue.items():
                service_stats.append({
                    'Dịch vụ': service_name,
                    'Số lần sử dụng': stats['count'],
                    'Tổng doanh thu (VNĐ)': f"{stats['total']:,}",
                    'Doanh thu trung bình': f"{stats['total'] // stats['count']:,}"
                })
            
            if service_stats:
                service_stats.sort(key=lambda x: int(x['Tổng doanh thu (VNĐ)'].replace(',', '')), reverse=True)
                data_dict['Thống kê theo dịch vụ'] = service_stats
            
            filename = self.generate_filename("DoanhThu", "xlsx", "BaoCao")
            filepath = self.get_report_path("thong_ke", filename)
            
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            result_path = self.create_summary_report(
                data_dict, 
                filepath, 
                "BÁO CÁO DOANH THU HỆ THỐNG", 
                author
            )
            
            return result_path
            
        except Exception as e:
            error(f"Lỗi khi xuất báo cáo doanh thu: {e}")
            return ""

def format_benh_nhan_data(benh_nhan_list: List) -> List[Dict]:
    data = []
    for i, bn in enumerate(benh_nhan_list, 1):
        data.append({
            'STT': i,
            'Mã BN': getattr(bn, 'ma_bn', ''),
            'PID': getattr(bn, 'pid', ''),
            'Họ tên': getattr(bn, '_ho_ten', ''),
            'Giới tính': getattr(bn, '_gioi_tinh', ''),
            'Năm sinh': getattr(bn, '_nam_sinh', ''),
            'CCCD': getattr(bn, '_so_cccd', ''),
            'Địa chỉ': getattr(bn, '_dia_chi', '') if hasattr(bn, '_dia_chi') else ''
        })
    return data

def format_tiep_nhan_data(tiep_nhan_list: List) -> List[Dict]:
    data = []
    for i, tn in enumerate(tiep_nhan_list, 1):
        data.append({
            'STT': i,
            'Mã tiếp nhận': getattr(tn, '_ma_tn', ''),
            'Mã BN': getattr(getattr(tn, '_bn', None), 'ma_bn', '') if hasattr(tn, '_bn') else '',
            'PID': getattr(getattr(tn, '_bn', None), 'pid', '') if hasattr(tn, '_bn') else '',
            'Tên bệnh nhân': getattr(getattr(tn, '_bn', None), '_ho_ten', '') if hasattr(tn, '_bn') else '',
            'CCCD': getattr(getattr(tn, '_bn', None), '_so_cccd', '') if hasattr(tn, '_bn') else '',
            'Dịch vụ': getattr(getattr(tn, '_dv', None), '_ten_dv', 'N/A') if hasattr(tn, '_dv') else 'N/A',
            'Giá dịch vụ (VNĐ)': f"{getattr(getattr(tn, '_dv', None), '_gia', 0):,}" if hasattr(tn, '_dv') else '0',
            'Phòng khám': getattr(getattr(tn, '_pk', None), '_ten_phong', 'N/A') if hasattr(tn, '_pk') else 'N/A',
            'Lý do khám': getattr(tn, '_ly_do', ''),
            'Bác sĩ': getattr(getattr(tn, '_bs', None), 'ho_ten', 'N/A') if hasattr(tn, '_bs') else 'N/A'
        })
    return data

def format_dich_vu_data(dich_vu_list: List) -> List[Dict]:
    data = []
    for i, dv in enumerate(dich_vu_list, 1):
        data.append({
            'STT': i,
            'Mã dịch vụ': getattr(dv, '_ma_dv', ''),
            'Tên dịch vụ': getattr(dv, '_ten_dv', ''),
            'Giá (VNĐ)': getattr(dv, '_gia', 0),
            'Giá định dạng': f"{getattr(dv, '_gia', 0):,} VNĐ"
        })
    return data

def format_phong_kham_data(phong_kham_list: List) -> List[Dict]:
    data = []
    for i, pk in enumerate(phong_kham_list, 1):
        data.append({
            'STT': i,
            'Mã phòng': getattr(pk, '_ma_phong', ''),
            'Tên phòng': getattr(pk, '_ten_phong', ''),
            'Mô tả': getattr(pk, '_mo_ta', '') if hasattr(pk, '_mo_ta') else ''
        })
    return data

def format_bac_si_data(bac_si_list: List) -> List[Dict]:
    data = []
    for i, bs in enumerate(bac_si_list, 1):
        data.append({
            'STT': i,
            'Mã bác sĩ': getattr(bs, 'ma_bs', ''),
            'Họ tên': getattr(bs, 'ho_ten', ''),
            'Chuyên khoa': getattr(bs, 'chuyen_khoa', ''),
            'Số điện thoại': getattr(bs, 'so_dt', ''),
            'Email': getattr(bs, 'email', ''),
            'Phòng khám': getattr(getattr(bs, '_phong_kham', None), '_ten_phong', 'Chưa gán') if hasattr(bs, '_phong_kham') else 'Chưa gán'
        })
    return data